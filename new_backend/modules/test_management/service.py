"""
test_management/service.py
─────────────────────────────
All business logic for the Test Management module — CRUD for the 9 entities,
the /run-tests orchestration, and dashboard aggregation. Flat "_flow"
functions, matching jira/service.py and test_runner/service.py: no
repository/controller layers, HTTPException raised directly.
"""

import json
import re
import uuid
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import HTTPException
from sqlalchemy import case, func, or_, select
from sqlalchemy.orm import Session

from new_backend.core.logger import logger
from new_backend.modules.test_management import models as schemas
from new_backend.modules.test_management.db_models import (
    Application,
    Attachment,
    Bug,
    ExecutionLog,
    Module,
    Priority,
    TestCase,
    TestRun,
    TestRunResult,
)
from new_backend.modules.test_management.runner import runner_service
from new_backend.modules.test_management.discovery import discover_automation_tests, normalize_match_key


# ─────────────────────────────────────────────────────────────────────────
# Small local helpers
# ─────────────────────────────────────────────────────────────────────────
def _paginate(page: int, page_size: int) -> tuple[int, int]:
    page = max(page, 1)
    page_size = min(max(page_size, 1), schemas.MAX_PAGE_SIZE)
    return page, page_size


def _paginated(items, total: int, page: int, page_size: int, schema) -> dict:
    total_pages = (total + page_size - 1) // page_size if page_size else 0
    return {
        "items": [schema.model_validate(i) for i in items],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


def _abbreviate(name: str, length: int) -> str:
    words = re.findall(r"[A-Za-z0-9]+", name or "")
    if not words:
        return "GEN"
    if len(words) == 1:
        return words[0][:length].upper()
    return "".join(w[0] for w in words)[:length].upper() or words[0][:length].upper()


def _next_testcase_key(db: Session, application_name: str, module_name: str) -> str:
    prefix = f"{_abbreviate(application_name, 3)}_{_abbreviate(module_name, 4)}"
    existing = db.scalars(select(TestCase.testcase_key).where(TestCase.testcase_key.like(f"{prefix}_%"))).all()
    max_seq = 0
    pattern = re.compile(rf"^{re.escape(prefix)}_(\d+)$")
    for key in existing:
        m = pattern.match(key)
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return f"{prefix}_{max_seq + 1:03d}"


# ═══════════════════════════════════════════════════════════════════════
# Applications
# ═══════════════════════════════════════════════════════════════════════
def create_application_flow(payload: schemas.ApplicationCreate, db: Session) -> Application:
    obj = Application(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_applications_flow(status: bool | None, q: str | None, page: int, page_size: int, db: Session) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if status is not None:
        filters.append(Application.status == status)
    if q:
        filters.append(Application.application_name.ilike(f"%{q}%"))

    stmt = select(Application).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(Application.application_name.asc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.ApplicationRead)


def get_application_flow(application_id: int, db: Session) -> Application:
    obj = db.get(Application, application_id)
    if not obj:
        raise HTTPException(404, f"Application {application_id} not found")
    return obj


def update_application_flow(application_id: int, payload: schemas.ApplicationUpdate, db: Session) -> Application:
    obj = get_application_flow(application_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_application_flow(application_id: int, db: Session) -> None:
    obj = get_application_flow(application_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Modules
# ═══════════════════════════════════════════════════════════════════════
def create_module_flow(payload: schemas.ModuleCreate, db: Session) -> Module:
    obj = Module(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_modules_flow(
    application_id: int | None, status: bool | None, page: int, page_size: int, db: Session
) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if application_id:
        filters.append(Module.application_id == application_id)
    if status is not None:
        filters.append(Module.status == status)

    stmt = select(Module).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(Module.display_order.asc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.ModuleRead)


def get_module_flow(module_id: int, db: Session) -> Module:
    obj = db.get(Module, module_id)
    if not obj:
        raise HTTPException(404, f"Module {module_id} not found")
    return obj


def update_module_flow(module_id: int, payload: schemas.ModuleUpdate, db: Session) -> Module:
    obj = get_module_flow(module_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_module_flow(module_id: int, db: Session) -> None:
    obj = get_module_flow(module_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Priorities
# ═══════════════════════════════════════════════════════════════════════
def create_priority_flow(payload: schemas.PriorityCreate, db: Session) -> Priority:
    obj = Priority(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_priorities_flow(page: int, page_size: int, db: Session) -> dict:
    page, page_size = _paginate(page, page_size)
    stmt = select(Priority)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(Priority.display_order.asc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.PriorityRead)


def get_priority_flow(priority_id: int, db: Session) -> Priority:
    obj = db.get(Priority, priority_id)
    if not obj:
        raise HTTPException(404, f"Priority {priority_id} not found")
    return obj


def update_priority_flow(priority_id: int, payload: schemas.PriorityUpdate, db: Session) -> Priority:
    obj = get_priority_flow(priority_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_priority_flow(priority_id: int, db: Session) -> None:
    obj = get_priority_flow(priority_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Test Cases
# ═══════════════════════════════════════════════════════════════════════
def create_test_case_flow(payload: schemas.TestCaseCreate, db: Session) -> TestCase:
    application = db.get(Application, payload.application_id)
    if not application:
        raise HTTPException(404, f"Application {payload.application_id} not found")
    module = db.get(Module, payload.module_id)
    if not module:
        raise HTTPException(404, f"Module {payload.module_id} not found")
    if module.application_id != payload.application_id:
        raise HTTPException(409, "Module does not belong to the given application")

    data = payload.model_dump()
    testcase_key = data.pop("testcase_key", None)
    if testcase_key:
        if db.scalar(select(TestCase).where(TestCase.testcase_key == testcase_key)):
            raise HTTPException(409, f"Test case key '{testcase_key}' already exists")
    else:
        testcase_key = _next_testcase_key(db, application.application_name, module.module_name)

    obj = TestCase(**data, testcase_key=testcase_key)
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


# ── Bulk upload from Excel ────────────────────────────────────────────────
# Header name (normalized: lowercased, non-alnum stripped) -> logical field.
_BULK_HEADER_ALIASES = {
    "testcaseid": "key",
    "testcasename": "title",
    "title": "title",
    "submodule": "sub_module",
    "testtype": "polarity",
    "priority": "priority",
    "testdata": "test_data",
    "preconditions": "preconditions",
    "teststeps": "test_steps",
    "expectedresult": "expected_result",
    "remarks": "remarks",
}


def _norm_header(h) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h or "").strip().lower())


# Canonical header row for the downloadable template (order matters for readability).
_BULK_TEMPLATE_HEADERS = [
    "Test Case ID", "Sub-Module", "Test Case Name", "Test Type", "Priority",
    "Test Data", "Preconditions", "Test Steps", "Expected Result", "remarks",
]


def build_bulk_template_xlsx() -> bytes:
    """Return an .xlsx (headers + one example row) users can fill in and re-upload."""
    import io

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server (pip install openpyxl).")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.append(_BULK_TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append([
        "RF_ONB_001",
        "Add Farm",
        "Add new farm with valid details",
        "Positive",                       # Test Type: Positive / Negative
        "High",                           # Priority: Critical / High / Medium / Low
        "Farm Name: Green Valley; Acreage: 3.5; Land Unit: Acre",
        "1. Farmer is logged in.\n2. The active farm list is displayed.",
        "1. Tap 'Add Farm'.\n2. Enter Farm Name and Acreage.\n3. Tap 'Submit'.",
        "Farm is created and appears in the farm list.",
        "Optional notes",
    ])
    ws.freeze_panes = "A2"
    for i, header in enumerate(_BULK_TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(header) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _compose_bulk_description(vals: dict) -> str | None:
    """Fold the Excel's rich columns into the single `description` field."""
    parts = []
    if vals.get("sub_module"):
        parts.append(f"Sub-Module: {vals['sub_module']}")
    if vals.get("test_data"):
        parts.append(f"Test Data:\n{vals['test_data']}")
    if vals.get("preconditions"):
        parts.append(f"Preconditions:\n{vals['preconditions']}")
    if vals.get("test_steps"):
        parts.append(f"Test Steps:\n{vals['test_steps']}")
    if vals.get("remarks"):
        parts.append(f"Remarks:\n{vals['remarks']}")
    return "\n\n".join(parts) or None


def bulk_upload_test_cases_flow(
    file_bytes: bytes, filename: str, application_id: int, module_id: int, db: Session
) -> dict:
    """
    Parse an .xlsx of test cases and create them under (application_id, module_id).
    Columns are matched by header name (see _BULK_HEADER_ALIASES); rows whose
    testcase_key already exists (in the DB or earlier in the same file) are skipped,
    not overwritten. Returns a per-file summary.
    """
    import io

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server (pip install openpyxl).")

    application = db.get(Application, application_id)
    if not application:
        raise HTTPException(404, f"Application {application_id} not found")
    module = db.get(Module, module_id)
    if not module:
        raise HTTPException(404, f"Module {module_id} not found")
    if module.application_id != application_id:
        raise HTTPException(409, "Module does not belong to the given application")

    logger.info(
        "[bulk-upload] Start | file=%s | app=%s | module=%s",
        filename, application.application_name, module.module_name,
    )

    if not (filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload an .xlsx file.")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read the Excel file: {e}")
    ws = wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(400, "The sheet is empty.")

    col_map: dict[int, str] = {}
    for idx, raw in enumerate(header_row):
        field = _BULK_HEADER_ALIASES.get(_norm_header(raw))
        if field and idx not in col_map:
            col_map[idx] = field
    if "title" not in col_map.values():
        raise HTTPException(
            400,
            "Could not find a 'Test Case Name' (title) column. Expected headers like: "
            "Test Case ID, Sub-Module, Test Case Name, Test Type, Priority, Test Data, "
            "Preconditions, Test Steps, Expected Result, remarks.",
        )
    logger.info("[bulk-upload] Columns detected: %s", sorted(set(col_map.values())))

    prio_by_name = {
        (p.priority_name or "").strip().lower(): p.priority_id
        for p in db.scalars(select(Priority)).all()
    }
    existing_keys = set(db.scalars(select(TestCase.testcase_key)).all())

    total_rows = created = 0
    created_keys: list[str] = []
    skipped_keys: list[str] = []
    error_details: list[str] = []
    seen_keys_this_file: set[str] = set()

    for i, row in enumerate(rows, start=2):  # 1-based, header was row 1
        vals: dict[str, str] = {}
        for idx, field in col_map.items():
            cell = row[idx] if idx < len(row) else None
            if cell is not None and str(cell).strip():
                vals[field] = str(cell).strip()

        if not vals:  # fully blank row
            continue
        total_rows += 1

        title = vals.get("title")
        if not title:
            error_details.append(f"Row {i}: missing Test Case Name — skipped.")
            continue

        key = vals.get("key") or None
        if key:
            if key in existing_keys or key in seen_keys_this_file:
                skipped_keys.append(key)
                logger.info("[bulk-upload] Row %d: skipped '%s' (already exists)", i, key)
                continue
            if len(key) > 50:
                error_details.append(f"Row {i}: Test Case ID '{key}' exceeds 50 chars — skipped.")
                logger.warning("[bulk-upload] Row %d: '%s' key exceeds 50 chars — skipped", i, key)
                continue

        polarity_raw = (vals.get("polarity") or "").strip().lower()
        polarity = "Positive" if polarity_raw.startswith("pos") else "Negative" if polarity_raw.startswith("neg") else None
        priority_id = prio_by_name.get((vals.get("priority") or "").strip().lower())

        resolved_key = key or _next_testcase_key(db, application.application_name, module.module_name)
        try:
            # SAVEPOINT per row: a bad row rolls back only itself, leaving every
            # previously-inserted row intact for the final commit (via get_db).
            with db.begin_nested():
                db.add(TestCase(
                    testcase_key=resolved_key,
                    title=title[:300],
                    application_id=application_id,
                    module_id=module_id,
                    priority_id=priority_id,
                    test_types=[],
                    polarity=polarity,
                    description=_compose_bulk_description(vals),
                    expected_result=vals.get("expected_result"),
                ))
        except Exception as e:
            msg = str(e).splitlines()[0][:200]
            error_details.append(f"Row {i} ({resolved_key}): {msg}")
            logger.error("[bulk-upload] Row %d: FAILED '%s' — %s", i, resolved_key, msg)
            continue

        created += 1
        created_keys.append(resolved_key)
        existing_keys.add(resolved_key)
        seen_keys_this_file.add(resolved_key)
        logger.info("[bulk-upload] Row %d: created '%s' — %s", i, resolved_key, title[:60])

    logger.info(
        "[bulk-upload] Done | file=%s | created=%d | skipped=%d | errors=%d | of %d rows",
        filename, created, len(skipped_keys), len(error_details), total_rows,
    )

    return {
        "total_rows": total_rows,
        "created": created,
        "skipped": len(skipped_keys),
        "errors": len(error_details),
        "created_keys": created_keys,
        "skipped_keys": skipped_keys,
        "error_details": error_details,
    }


def list_test_cases_flow(
    application_id: int | None,
    module_id: int | None,
    priority_id: int | None,
    test_type: str | None,
    status: bool | None,
    polarity: str | None,
    q: str | None,
    page: int,
    page_size: int,
    db: Session,
) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if application_id:
        filters.append(TestCase.application_id == application_id)
    if module_id:
        filters.append(TestCase.module_id == module_id)
    if priority_id:
        filters.append(TestCase.priority_id == priority_id)
    if status is not None:
        filters.append(TestCase.status == status)
    if polarity:
        filters.append(TestCase.polarity == polarity)
    if test_type:
        filters.append(func.json_contains(TestCase.test_types, json.dumps(test_type)))
    if q:
        like = f"%{q}%"
        filters.append(or_(TestCase.title.ilike(like), TestCase.testcase_key.ilike(like), TestCase.description.ilike(like)))

    stmt = select(TestCase).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(TestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.TestCaseRead)


def get_test_case_flow(testcase_id: int, db: Session) -> TestCase:
    obj = db.get(TestCase, testcase_id)
    if not obj:
        raise HTTPException(404, f"Test case {testcase_id} not found")
    return obj


def find_test_case_by_source_name(
    name: str, db: Session, application_id: int | None = None, module_id: int | None = None
) -> TestCase | None:
    """
    Resolve a running/source test name (e.g. the pytest function `test_LOGINPOS_TC_029`)
    to its catalogued TestCase by comparing on `normalize_match_key` — so a
    `test_`/`tc_` prefix in code matches a sheet/DB key that omits it
    (`LOGINPOS_TC_029`). Optionally scope to an application/module to disambiguate
    keys that repeat across suites. Used to store a test result against the right
    test case. Returns None when nothing matches.
    """
    target = normalize_match_key(name)
    if not target:
        return None
    filters = []
    if application_id is not None:
        filters.append(TestCase.application_id == application_id)
    if module_id is not None:
        filters.append(TestCase.module_id == module_id)
    for tc in db.scalars(select(TestCase).where(*filters)):
        if normalize_match_key(tc.testcase_key) == target:
            return tc
    return None


# ── Live pytest run reporting ─────────────────────────────────────────────
# Maps a pytest outcome to the TestResultStatus vocabulary.
_PYTEST_STATUS_MAP = {"passed": "PASSED", "failed": "FAILED", "skipped": "SKIPPED", "error": "FAILED"}


def start_pytest_run_flow(payload: schemas.PytestRunStart, db: Session) -> TestRun:
    """Create a RUNNING TestRun for a real Appium/pytest execution. The automation
    variant (app_type) is resolved to the DB application_id so results can be scoped."""
    application_id = None
    if payload.app_type:
        app = db.scalars(select(Application).where(Application.variant == payload.app_type)).first()
        application_id = app.application_id if app else None
    run = TestRun(
        run_name=payload.run_name or f"Automation {datetime.utcnow():%Y-%m-%d %H:%M:%S}",
        application_id=application_id,
        environment=payload.environment,
        build_number=payload.build_number,
        execution_type="Automated",
        triggered_by=payload.triggered_by,
        status="RUNNING",
        started_at=datetime.utcnow(),
    )
    db.add(run)
    db.flush()
    db.refresh(run)
    logger.info("[test-run] started run_id=%s app_id=%s (app_type=%s)", run.run_id, application_id, payload.app_type)
    return run


def record_pytest_result_flow(run_id: int, payload: schemas.PytestResultRecord, db: Session) -> TestRunResult:
    """Store one test's outcome, matched to its catalogued TestCase by name
    (test_/tc_ prefix stripped). Unmatched tests are still recorded with a null
    testcase_id so nothing is lost."""
    run = db.get(TestRun, run_id)
    if not run:
        raise HTTPException(404, f"Test run {run_id} not found")

    tc = find_test_case_by_source_name(payload.test_name, db, application_id=run.application_id)
    status = _PYTEST_STATUS_MAP.get((payload.status or "").strip().lower(), (payload.status or "").upper())
    result = TestRunResult(
        run_id=run_id,
        testcase_id=tc.testcase_id if tc else None,
        status=status,
        execution_time=Decimal(str(round(payload.execution_time, 2))) if payload.execution_time is not None else None,
        device_name=payload.device_name,
        os_version=payload.os_version,
        browser=payload.browser,
        failure_reason=payload.failure_reason,
        completed_at=datetime.utcnow(),
    )
    db.add(result)
    db.flush()
    db.refresh(result)
    logger.info(
        "[test-run] run=%s test=%s -> testcase=%s status=%s",
        run_id, payload.test_name, (tc.testcase_key if tc else "UNMATCHED"), status,
    )
    return result


def finish_pytest_run_flow(run_id: int, payload: schemas.PytestRunFinish, db: Session) -> TestRun:
    """Finalize a run: explicit status if given, else FAILED when any result failed."""
    run = db.get(TestRun, run_id)
    if not run:
        raise HTTPException(404, f"Test run {run_id} not found")
    if payload.status:
        run.status = payload.status
    else:
        statuses = db.scalars(select(TestRunResult.status).where(TestRunResult.run_id == run_id)).all()
        run.status = "FAILED" if any((s or "").upper() in ("FAILED", "ERROR", "BLOCKED") for s in statuses) else "COMPLETED"
    run.completed_at = datetime.utcnow()
    db.flush()
    db.refresh(run)
    logger.info("[test-run] finished run_id=%s status=%s", run_id, run.status)
    return run


def update_test_case_flow(testcase_id: int, payload: schemas.TestCaseUpdate, db: Session) -> TestCase:
    obj = get_test_case_flow(testcase_id, db)
    data = payload.model_dump(exclude_unset=True)
    if "module_id" in data:
        module = db.get(Module, data["module_id"])
        if not module:
            raise HTTPException(404, f"Module {data['module_id']} not found")
        if module.application_id != obj.application_id:
            raise HTTPException(409, "Module does not belong to this test case's application")
    for field, value in data.items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_test_case_flow(testcase_id: int, db: Session) -> None:
    obj = get_test_case_flow(testcase_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Test Runs — POST /run-tests orchestration
# ═══════════════════════════════════════════════════════════════════════
def run_tests_flow(payload: schemas.TestRunCreate, db: Session) -> dict:
    started = datetime.utcnow()

    run = TestRun(
        run_name=payload.run_name or f"Run {started.strftime('%Y-%m-%d %H:%M:%S')}",
        application_id=payload.application_id,
        module_id=payload.module_id,
        test_type=payload.test_type,
        environment=payload.environment,
        build_number=payload.build_number,
        execution_type=payload.execution_type,
        triggered_by=payload.triggered_by,
        status="RUNNING",
        started_at=started,
    )
    db.add(run)
    db.flush()
    db.refresh(run)

    filters = [TestCase.application_id == payload.application_id, TestCase.status.is_(True)]
    if payload.module_id:
        filters.append(TestCase.module_id == payload.module_id)
    if payload.test_type:
        filters.append(func.json_contains(TestCase.test_types, json.dumps(payload.test_type)))

    matching_cases = db.scalars(select(TestCase).where(*filters)).all()
    testcase_ids = [tc.testcase_id for tc in matching_cases]

    # Still a valid, informative outcome when nothing matches — not an error.
    runner_results = (
        runner_service.run_tests(testcase_ids, payload.environment, payload.execution_type) if testcase_ids else []
    )

    created_results: list[TestRunResult] = []
    for r in runner_results:
        result = TestRunResult(
            run_id=run.run_id,
            testcase_id=r.testcase_id,
            status=r.status,
            execution_time=Decimal(str(r.execution_time)),
            device_name=r.device_name,
            os_version=r.os_version,
            browser=r.browser,
            failure_reason=r.failure_reason,
            allure_report=None,
            started_at=r.started_at,
            completed_at=r.completed_at,
        )
        db.add(result)
        created_results.append(result)
    if created_results:
        db.flush()
        for result in created_results:
            db.refresh(result)

    log_rows = [
        ExecutionLog(execution_id=db_result.execution_id, log_level=level, message=message)
        for db_result, runner_result in zip(created_results, runner_results)
        for level, message in runner_result.log_lines
    ]
    if log_rows:
        db.add_all(log_rows)
        db.flush()

    completed = datetime.utcnow()
    passed = sum(1 for r in runner_results if r.status == "PASSED")
    failed = sum(1 for r in runner_results if r.status == "FAILED")
    skipped = len(runner_results) - passed - failed

    run.status = "COMPLETED"
    run.completed_at = completed
    db.flush()
    db.refresh(run)

    logger.info(
        "run-tests: run_id=%s total=%s passed=%s failed=%s skipped=%s",
        run.run_id, len(runner_results), passed, failed, skipped,
    )

    return {
        "run": schemas.TestRunRead.model_validate(run),
        "total": len(runner_results),
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "results": [schemas.TestRunResultRead.model_validate(r) for r in created_results],
    }


def discover_automation_tests_flow(path: str) -> list[dict]:
    return discover_automation_tests(path)


def list_test_runs_flow(
    application_id: int | None, module_id: int | None, status: str | None, page: int, page_size: int, db: Session
) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if application_id:
        filters.append(TestRun.application_id == application_id)
    if module_id:
        filters.append(TestRun.module_id == module_id)
    if status:
        filters.append(TestRun.status == status)

    stmt = select(TestRun).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(TestRun.started_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.TestRunRead)


def get_test_run_flow(run_id: int, db: Session) -> TestRun:
    obj = db.get(TestRun, run_id)
    if not obj:
        raise HTTPException(404, f"Test run {run_id} not found")
    return obj


def cancel_test_run_flow(run_id: int, db: Session) -> TestRun:
    obj = get_test_run_flow(run_id, db)
    runner_service.cancel_execution(run_id)
    obj.status = "CANCELLED"
    obj.completed_at = datetime.utcnow()
    db.flush()
    db.refresh(obj)
    return obj


# ═══════════════════════════════════════════════════════════════════════
# Test Run Results
# ═══════════════════════════════════════════════════════════════════════
def create_test_run_result_flow(payload: schemas.TestRunResultCreate, db: Session) -> TestRunResult:
    obj = TestRunResult(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_test_run_results_flow(
    run_id: int | None, testcase_id: int | None, status: str | None, page: int, page_size: int, db: Session
) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if run_id:
        filters.append(TestRunResult.run_id == run_id)
    if testcase_id:
        filters.append(TestRunResult.testcase_id == testcase_id)
    if status:
        filters.append(TestRunResult.status == status)

    stmt = select(TestRunResult).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(TestRunResult.started_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.TestRunResultRead)


def get_test_run_result_flow(execution_id: int, db: Session) -> TestRunResult:
    obj = db.get(TestRunResult, execution_id)
    if not obj:
        raise HTTPException(404, f"Execution result {execution_id} not found")
    return obj


def update_test_run_result_flow(execution_id: int, payload: schemas.TestRunResultUpdate, db: Session) -> TestRunResult:
    obj = get_test_run_result_flow(execution_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_test_run_result_flow(execution_id: int, db: Session) -> None:
    obj = get_test_run_result_flow(execution_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Execution Logs — append-only
# ═══════════════════════════════════════════════════════════════════════
def create_execution_logs_flow(payload: schemas.ExecutionLogBulkCreate, db: Session) -> list[ExecutionLog]:
    if not payload.logs:
        return []
    objs = [
        ExecutionLog(execution_id=payload.execution_id, log_level=entry.log_level, message=entry.message)
        for entry in payload.logs
    ]
    db.add_all(objs)
    db.flush()
    for obj in objs:
        db.refresh(obj)
    return objs


def list_execution_logs_flow(execution_id: int, db: Session) -> list[ExecutionLog]:
    return db.scalars(
        select(ExecutionLog).where(ExecutionLog.execution_id == execution_id).order_by(ExecutionLog.created_at.asc())
    ).all()


# ═══════════════════════════════════════════════════════════════════════
# Attachments
# ═══════════════════════════════════════════════════════════════════════
def create_attachment_flow(payload: schemas.AttachmentCreate, db: Session) -> Attachment:
    obj = Attachment(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_attachments_flow(execution_id: int | None, page: int, page_size: int, db: Session) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if execution_id:
        filters.append(Attachment.execution_id == execution_id)

    stmt = select(Attachment).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(Attachment.uploaded_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.AttachmentRead)


def get_attachment_flow(attachment_id: int, db: Session) -> Attachment:
    obj = db.get(Attachment, attachment_id)
    if not obj:
        raise HTTPException(404, f"Attachment {attachment_id} not found")
    return obj


def update_attachment_flow(attachment_id: int, payload: schemas.AttachmentUpdate, db: Session) -> Attachment:
    obj = get_attachment_flow(attachment_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(obj, field, value)
    db.flush()
    db.refresh(obj)
    return obj


def delete_attachment_flow(attachment_id: int, db: Session) -> None:
    obj = get_attachment_flow(attachment_id, db)
    db.delete(obj)
    db.flush()


# ═══════════════════════════════════════════════════════════════════════
# Bugs
# ═══════════════════════════════════════════════════════════════════════
def create_bug_flow(payload: schemas.BugCreate, db: Session) -> Bug:
    obj = Bug(**payload.model_dump())
    db.add(obj)
    db.flush()
    db.refresh(obj)
    return obj


def list_bugs_flow(
    testcase_id: int | None,
    execution_id: int | None,
    status: str | None,
    severity: str | None,
    page: int,
    page_size: int,
    db: Session,
) -> dict:
    page, page_size = _paginate(page, page_size)
    filters = []
    if testcase_id:
        filters.append(Bug.testcase_id == testcase_id)
    if execution_id:
        filters.append(Bug.execution_id == execution_id)
    if status:
        filters.append(Bug.status == status)
    if severity:
        filters.append(Bug.severity == severity)

    stmt = select(Bug).where(*filters)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    items = db.scalars(
        stmt.order_by(Bug.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return _paginated(items, total, page, page_size, schemas.BugRead)


def get_bug_flow(bug_id: int, db: Session) -> Bug:
    obj = db.get(Bug, bug_id)
    if not obj:
        raise HTTPException(404, f"Bug {bug_id} not found")
    return obj


# ═══════════════════════════════════════════════════════════════════════
# Dashboard
# ═══════════════════════════════════════════════════════════════════════
def dashboard_summary_flow(db: Session) -> dict:
    total_testcases = db.scalar(select(func.count()).select_from(TestCase)) or 0
    smoke_count = db.scalar(
        select(func.count()).select_from(TestCase).where(func.json_contains(TestCase.test_types, json.dumps("Smoke")))
    ) or 0
    regression_count = db.scalar(
        select(func.count()).select_from(TestCase).where(func.json_contains(TestCase.test_types, json.dumps("Regression")))
    ) or 0

    total_results = db.scalar(select(func.count()).select_from(TestRunResult)) or 0
    passed_results = db.scalar(select(func.count()).select_from(TestRunResult).where(TestRunResult.status == "PASSED")) or 0
    failed_results = db.scalar(select(func.count()).select_from(TestRunResult).where(TestRunResult.status == "FAILED")) or 0

    pass_percentage = round((passed_results / total_results) * 100, 2) if total_results else 0.0
    fail_percentage = round((failed_results / total_results) * 100, 2) if total_results else 0.0

    latest_runs = db.scalars(select(TestRun).order_by(TestRun.started_at.desc()).limit(5)).all()
    recent_failures = db.scalars(
        select(TestRunResult).where(TestRunResult.status == "FAILED").order_by(TestRunResult.completed_at.desc()).limit(5)
    ).all()

    return {
        "total_testcases": total_testcases,
        "smoke_count": smoke_count,
        "regression_count": regression_count,
        "pass_percentage": pass_percentage,
        "fail_percentage": fail_percentage,
        "latest_runs": [schemas.LatestRun.model_validate(r) for r in latest_runs],
        "recent_failures": [schemas.RecentFailure.model_validate(r) for r in recent_failures],
    }


def execution_history_flow(days: int, db: Session) -> list[dict]:
    since = datetime.utcnow() - timedelta(days=days)
    day_col = func.date(TestRunResult.completed_at)
    passed_expr = func.sum(case((TestRunResult.status == "PASSED", 1), else_=0))
    failed_expr = func.sum(case((TestRunResult.status == "FAILED", 1), else_=0))

    rows = db.execute(
        select(day_col.label("day"), passed_expr.label("passed"), failed_expr.label("failed"), func.count().label("total"))
        .where(TestRunResult.completed_at.isnot(None), TestRunResult.completed_at >= since)
        .group_by(day_col)
        .order_by(day_col.asc())
    ).all()

    return [
        schemas.ExecutionHistoryPoint(date=row.day, passed=row.passed or 0, failed=row.failed or 0, total=row.total or 0)
        for row in rows
    ]
